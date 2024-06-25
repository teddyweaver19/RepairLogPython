import openpyxl 
from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import tkinter as tk
import openpyxl, xdrlib
from openpyxl import Workbook 
import pathlib 
from tkinter.filedialog import *
from hello import *
from datetime import datetime

#Initialize the form 
root = Tk()
root.title("Automated Repair Log") #Title of the form
root.geometry('700x400+300+200') #Size of the form 
root.resizable(False,False) #Option to make it resizable 
root.configure(bg="#326273") #Set the background color 

file = pathlib.Path('RepairLog.xlsx')

if file.exists(): #If the file exists in the directory, pass, else create a new workbook and use formatting from 'hello.py' to freeze panes, set column width, row height, and labels
    pass
else:
    wb = Workbook()
    ws = wb.active
    freeze_pane(wb)
    text_format(ws)
    wb.save('RepairLog.xlsx')
    


def clear(): #Clears the form and resets the ComboBoxes 
    IssueVal.set('')
    EquipID.set('')
    IssueType_Combo.set('Select a type')
    PriorityLvl_Combo.set('Select a priority level')
    Hours.set('')
    Called_In.set('')
    AssignedTech.set('')

def submit(): #Submits the form after retrieving the data entered into the form 
    Equip = EquipID.get()
    Issue = IssueVal.get()
    IssueType = IssueType_Combo.get()
    Priority = PriorityLvl_Combo.get()
    Hrs = Hours.get()
    Called = Called_In.get()
    Assigned = AssignedTech.get()

    timestamp = datetime.now()

    wb = openpyxl.load_workbook('RepairLog.xlsx') #Open the workbook
    ws = wb.active #Select the sheet

    #Assign values to the cells in the last row without info (max_row + 1 and then max_row to avoid writing over headers)
    ws.cell(column=1,row=ws.max_row+1,value=Equip)
    ws.cell(column=2,row=ws.max_row,value=Issue)
    ws.cell(column=3,row=ws.max_row,value=IssueType)
    ws.cell(column=4,row=ws.max_row,value=Priority)
    ws.cell(column=5, row=ws.max_row,value=timestamp)
    ws.cell(column=7,row=ws.max_row,value=Assigned)
    ws.cell(column=8,row=ws.max_row,value=Called)
    ws.cell(column=10,row=ws.max_row,value=Hrs)

    wb.save(r'RepairLog.xlsx')

    messagebox.showinfo('Complete','Entry completed')

#Icon
icon_image = PhotoImage(file ="SSA.png") #Set Icon photo
root.iconphoto(False,icon_image)

#Header 
Label(root,text="Please fill out this repair log:", font="arial 13",bg="#326273",fg="#fff" ).place(x=20,y=20) #Creates a heading label and puts it at x=20 y = 20 (top left)


#Labels 
Label(root,text='Equipment ID:', font="arial 13",bg="#326273",fg="#fff").place(x=50,y=100) #Creates a label at (50,100)
Label(root,text='Issue:', font="arial 13",bg="#326273",fg="#fff").place(x=50,y=130) #Creates a label at (50,130)
Label(root,text='Type:', font="arial 13",bg="#326273",fg="#fff").place(x=50,y=160) #Creates a label at (50,160)
Label(root,text='Priority Level:', font="arial 13",bg="#326273",fg="#fff").place(x=50,y=190) #Creates a label at (50,190)
Label(root,text='Assigned Tech:', font="arial 13",bg="#326273",fg="#fff").place(x=50,y=220) #Creates a label at (50,220)
Label(root,text='Called in by:', font="arial 13",bg="#326273",fg="#fff").place(x=50,y=250) #Creates a label at (50,250)
Label(root,text='Hours:', font="arial 13",bg="#326273",fg="#fff").place(x=50,y=280) #Creates a label at (50,280)

#Entries

#IssueVal Entry
IssueVal = StringVar()
IssueVal_Entry = Entry(root,textvariable=IssueVal, width=20,bd=2,font=20)
IssueVal_Entry.place(x=170, y=130)

#IssueType ComboBox
IssueType_Combo = Combobox(root,values=['Select a type','Repair-Field','Repair-Warranty','Repair-Shop','Repair-Damage','Repair-GPS','Operations','PM'],font='arial 12',state='r', width=20)
IssueType_Combo.place(x=170,y=160)
IssueType_Combo.set('Select a type') #Default selection upon opening form

#EquipmentID Entry
EquipID = StringVar()
EquipID_Entry = Entry(root,textvariable=EquipID, width=20,bd=2,font=20)
EquipID_Entry.place(x=170, y=100)


#PriorityLvl ComboBox
PriorityLvl_Combo = Combobox(root,values=['Select a priority level','Low','Medium','High'],font='arial 12',state='r', width=20)
PriorityLvl_Combo.place(x=170,y=190)
PriorityLvl_Combo.set('Select a priority level') #Default selection upon opening form

#AssignedTech Entry 
AssignedTech = StringVar()
AssignedTech_Entry = Entry(root,textvariable=AssignedTech,width=20,bd=2,font=20)
AssignedTech_Entry.place(x=170,y=220)

#CalledIn Entry
Called_In = StringVar()
Called_In_Entry = Entry(root,textvariable=Called_In,width=20,bd=2,font=20)
Called_In_Entry.place(x=170,y=250)

#Hours Entry
Hours = StringVar()
Hours_Entry = Entry(root,textvariable=Hours,width=20,bd=2,font=20)
Hours_Entry.place(x=170,y=280)

#Buttons 
Button(root,text="Submit", bg ="#326723", fg="white", width=15,height=2,command=submit).place(x=150,y=350)
Button(root,text="Reset", bg ="#326723", fg="white", width=15,height=2,command=clear).place(x=300,y=350)
Button(root,text="Cancel", bg ="#326723", fg="white", width=15,height=2,command=lambda:root.destroy()).place(x=450,y=350) #Exits the form

root.mainloop()

