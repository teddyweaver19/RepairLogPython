import openpyxl
from openpyxl.styles import Font , PatternFill


wb = openpyxl.load_workbook('FormatPractice.xlsx')
ws = wb['Sheet1']

def freeze_pane(wb):
    sheet = wb.active
    sheet.freeze_panes = 'A2' #Freeze the top row 
    


def text_format(ws):

    ws['A1'] = '#'
    ws['B1'] = 'ISSUE'
    ws['C1'] = 'TYPE'
    ws['D1'] = 'PRIORITY'
    ws['E1'] = 'OPENED'
    ws['F1'] = 'CLOSED'
    ws['G1'] = 'ASSIGNED TECH'
    ws['H1'] = 'CALLED IN BY'
    ws['I1'] = 'NOTES'
    ws['J1'] = 'HOURS'
    ws['K1'] = 'STATUS'

#set row and column dimensions , they vary depending on expected informaton length  
    ws.row_dimensions[1].height = 35
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25 
    ws.column_dimensions['I'].width = 75
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 15
    
    for cell in ws["1:1"]: #For each cell in row 1 of the current worksheet
        cell.font = Font(name='Calibri', size=16, bold=True, italic=True, color='FFFFFF') #Change the font to white and set font type,size,bold,italics
        cell.fill = PatternFill(start_color='595959', end_color='595959', fill_type='solid') #Fill the entire row grey
    
freeze_pane(wb)
text_format(ws)

wb.save('FormatPractice2.xlsx')

print("Successful")